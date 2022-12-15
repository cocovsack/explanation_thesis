from collections import defaultdict
import numpy as np
import json
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# from prettytable import PrettyTable


dir_path = 'src/chefbot_utils/explainability_trials'

def main():
    workbook = Workbook()
    sheet = workbook.active
    row_count = 3

    action_overlay_list = []
    
    #Print ACTION count
    for i in range(3, 29):
        sheet.cell(row = 1, column = i).value = "ACTION: {}".format(i - 2)

    #Loop through jsons of each overlay combination
    for file in sorted(os.listdir(dir_path)):
        data = json.load(open(os.path.join(dir_path, file)))

        overlays = data['overlays']
        no_overlay_actions = data['predicted_actions']
        actual_actions = data['actual_actions']
        overlay_combo = data['id']

        if overlay_combo == ['1', '2', '3']:
            max_overlay_actions = actual_actions
        else:
            action_overlay_list.append([overlay_combo, actual_actions])

        #Print overlay arrays
        sheet.cell(row = row_count, column=1).value = str(overlay_combo)
        
        #Print overlay strings
        sheet.cell(row = row_count, column=2).value = ('___').join(overlays)

        #Print predicted actions
        sheet.cell(row=2, column=1).value = "[]"
        for col_count in range(3, len(no_overlay_actions) + 3):
            sheet.cell(row = 2, column = col_count).value = no_overlay_actions[col_count - 3]

        #Print row of actions for that overlay combo
        for col_count in range(3, len(actual_actions) + 3):
            sheet.cell(row = row_count, column = col_count).value = actual_actions[col_count - 3]

        row_count += 1
                    
    # ABLATION ON THE NO OVERLAY
    # for i in range(0, len(no_overlay_actions)):
    #     print(no_overlay_actions[i])
    #     for item in action_overlay_list:
    #         ids = item[0]
    #         actions = item[1]
    #         print("****ACTION: {}".format(no_overlay_actions[i]))
    #         if no_overlay_actions[i] == actions[i]:
    #             print("The same action for {}".format(ids))
    #         elif no_overlay_actions[i] in actions[0:i]:
    #             print("Already done for {}".format(ids))
    #         else:
    #             print("Not done by {}".format(ids))

    # Track which overlays id combinations have the same/similar action sequence
    exact_same = defaultdict(list)
    already_done = defaultdict(list)
    no_similarity = defaultdict(list)
    # Loop over each action when all overlays are used
    for i in range(0, len(max_overlay_actions)):
        # Loop over all the other overlay combinations and compare their actions
        for item in action_overlay_list:
            ids = item[0]
            actions = item[1]
            # Make sure not out of rane
            if i < len(actions):
                if max_overlay_actions[i] == actions[i]:
                    exact_same[i + 1].append(ids)
                elif max_overlay_actions[i] in actions[0:i]:
                    # print(actions.index(max_overlay_actions[i]))
                    already_done[actions.index(max_overlay_actions[i]) + 1].append(ids)
                else:
                    no_similarity[i + 1].append(ids)

    # EXACT SAME: Color the cells that are the exact same as the complete overlay
    for action_number in exact_same.keys():
        for i in range(len(exact_same.get(action_number))):
            overlays = exact_same[action_number][i]

            # Find which row the overlay combo is in
            row_var = 1
            while sheet.cell(row=row_var, column=1).value != str(overlays):
                row_var += 1
            
            # Color the cell to indicate they are the same
            sheet.cell(row = row_var, column = action_number + 2).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type = "solid")


    # ALMOST THE SAME: Color the cells that are almost the same as the complete overlay
    print(already_done)
    for action_number in already_done.keys():
        for i in range(len(already_done.get(action_number))):
            overlays = already_done[action_number][i]

            # Find which row the overlay combo is in
            row_var = 1
            while sheet.cell(row=row_var, column=1).value != str(overlays):
                row_var += 1
            
            # Color the cell to indicate they are the same
            sheet.cell(row = row_var, column = action_number + 2).fill = PatternFill(start_color="FFFF8F", end_color="FFFF8F", fill_type = "solid")

    workbook.save('/Users/cocosack/Main Folder/Yale \'22-\'23/Thesis/experiment3.xlsx')





            
if __name__ == "__main__":
    main()
